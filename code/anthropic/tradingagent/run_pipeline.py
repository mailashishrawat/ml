#!/usr/bin/env python3.14
"""
Indian Stock Market Recommendation Pipeline
Date: 2026-05-08
Orchestrates: Screener -> Analyzer -> Validator -> Formatter -> Backtest
"""

import anthropic
import json
import os
import sys
from datetime import datetime, date
from pathlib import Path

# ─── Configuration ───────────────────────────────────────────────────────────
WORKING_DIR = Path("/Users/I038849/Documents/Ashish/github.com/iimb/ml/code/anthropic/tradingagent")
TODAY = date(2026, 5, 8)
TODAY_STR = "2026-05-08"

BASESTOCK_FILE = WORKING_DIR / "basestock.xlsx"
PATTERN_NOTES  = WORKING_DIR / "pattern_notes.md"
DUOPOLY_FILE   = WORKING_DIR / "duopoly_pairs.json"
DAILY_REC_FILE = WORKING_DIR / "daily_recommendations.json"

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))

token_usage = {
    "screener": {"input": 0, "output": 0},
    "analyzer": {"input": 0, "output": 0},
    "validator": {"input": 0, "output": 0},
    "formatter": {"input": 0, "output": 0},
}

# ─── Helper ──────────────────────────────────────────────────────────────────
def load_file_text(path: Path) -> str:
    if path.exists():
        return path.read_text()
    return ""

def call_claude(role_name: str, model: str, system: str, user: str,
                max_tokens: int = 4096) -> str:
    resp = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        system=system,
        messages=[{"role": "user", "content": user}],
    )
    token_usage[role_name]["input"]  += resp.usage.input_tokens
    token_usage[role_name]["output"] += resp.usage.output_tokens
    return resp.content[0].text

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 – BASE STOCK SCREENER (Haiku)
# ═══════════════════════════════════════════════════════════════════════════════
def step1_screener() -> dict:
    """
    Because we cannot call live NSE/Chartink APIs in this environment,
    the Haiku sub-agent performs fundamental screening using its training
    knowledge of listed Indian mid/small-cap stocks (as of knowledge cutoff),
    applying the stated criteria logically and returning a filtered list.
    """
    print("\n=== STEP 1: BASE STOCK SCREENER (Haiku) ===")

    # Check if basestock data already exists for current month
    month_tag = TODAY.strftime("%Y-%m")
    cached_data = None
    if BASESTOCK_FILE.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BASESTOCK_FILE)
            meta_ws = wb["Metadata"] if "Metadata" in wb.sheetnames else None
            if meta_ws:
                gen_date_str = meta_ws["B1"].value
                if gen_date_str and str(gen_date_str)[:7] == month_tag:
                    print(f"  [CACHE HIT] basestock.xlsx is current ({gen_date_str}). Skipping re-screen.")
                    ws = wb.active
                    rows = []
                    headers = [c.value for c in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            rows.append(dict(zip(headers, row)))
                    cached_data = rows
        except Exception as e:
            print(f"  [WARN] Could not read cache: {e}")

    if cached_data:
        return {"stocks": cached_data, "from_cache": True}

    system_prompt = """You are a stock screener agent for the Indian equity market.
Your task is to identify mid-cap and small-cap NSE/BSE stocks that meet ALL of the following criteria:

a. Market cap classification: Mid Cap (₹5,000 Cr – ₹20,000 Cr) or Small Cap (₹500 Cr – ₹5,000 Cr)
b. Last closing stock price > ₹25
c. Market cap > ₹1,000 Crore
d. Year-over-year: Profit increment > 25% OR loss reduction > 25%
e. Average daily turnover (Close × Volume) > ₹1 Crore

Use your knowledge of Indian listed companies (NSE/BSE) as of early 2026. Focus on sectors with momentum:
EV, Solar/Renewable, Defense, Capital Goods, Chemicals, Pharma, Consumer Durables, IT Services, BFSI.

Return a JSON array of 25-35 stocks (no markdown fences). Each object must have:
{
  "symbol": "NSE symbol",
  "company_name": "Full name",
  "market_cap_cr": number,
  "last_close": number,
  "yoy_profit_change_pct": number,
  "avg_daily_turnover_lakh": number,
  "sector": "string",
  "industry": "string",
  "cap_category": "Mid Cap" or "Small Cap"
}

Return ONLY the JSON array, no other text."""

    user_msg = f"""Date: {TODAY_STR}
Screen for quality Indian mid-cap and small-cap stocks meeting ALL criteria.
Focus on companies with strong fundamentals, recent earnings momentum, and healthy liquidity.
Return 25-35 stocks as a JSON array."""

    print("  Calling Haiku for stock screening...")
    raw = call_claude("screener", "claude-haiku-4-5", system_prompt, user_msg, max_tokens=8000)

    # Parse JSON
    try:
        # Strip any accidental markdown
        raw_clean = raw.strip()
        if raw_clean.startswith("```"):
            raw_clean = raw_clean.split("```")[1]
            if raw_clean.startswith("json"):
                raw_clean = raw_clean[4:]
        stocks = json.loads(raw_clean)
    except json.JSONDecodeError as e:
        print(f"  [WARN] JSON parse error: {e}. Using fallback list.")
        stocks = _fallback_stocks()

    # Save to xlsx
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stocks"
        headers = ["Symbol", "Company Name", "Market Cap (Cr)", "Last Close (₹)",
                   "YoY Profit Change (%)", "Avg Daily Turnover (Lakh ₹)", "Sector", "Industry", "Cap Category"]
        ws.append(headers)
        for s in stocks:
            ws.append([
                s.get("symbol", ""),
                s.get("company_name", ""),
                s.get("market_cap_cr", 0),
                s.get("last_close", 0),
                s.get("yoy_profit_change_pct", 0),
                s.get("avg_daily_turnover_lakh", 0),
                s.get("sector", ""),
                s.get("industry", ""),
                s.get("cap_category", ""),
            ])
        meta_ws = wb.create_sheet("Metadata")
        meta_ws["A1"] = "Generated On"
        meta_ws["B1"] = TODAY_STR
        meta_ws["A2"] = "Stock Count"
        meta_ws["B2"] = len(stocks)
        wb.save(BASESTOCK_FILE)
        print(f"  Saved {len(stocks)} stocks to basestock.xlsx")
    except ImportError:
        print("  [WARN] openpyxl not available, skipping xlsx save")
    except Exception as e:
        print(f"  [WARN] Could not save xlsx: {e}")

    return {"stocks": stocks, "from_cache": False}


def _fallback_stocks():
    """Fallback if LLM fails to return valid JSON."""
    return [
        {"symbol": "IRCTC", "company_name": "Indian Railway Catering and Tourism Corp",
         "market_cap_cr": 12500, "last_close": 780, "yoy_profit_change_pct": 32,
         "avg_daily_turnover_lakh": 450, "sector": "Consumer Discretionary", "industry": "Railways Tourism", "cap_category": "Mid Cap"},
        {"symbol": "CDSL", "company_name": "Central Depository Services Ltd",
         "market_cap_cr": 8500, "last_close": 1420, "yoy_profit_change_pct": 45,
         "avg_daily_turnover_lakh": 380, "sector": "Financial Services", "industry": "Depository", "cap_category": "Mid Cap"},
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 – PATTERN ANALYSIS & RECOMMENDATION (Opus)
# ═══════════════════════════════════════════════════════════════════════════════
def step2_analyzer(stocks: list) -> list:
    print("\n=== STEP 2: PATTERN ANALYSIS & RECOMMENDATION (Opus) ===")

    pattern_notes = load_file_text(PATTERN_NOTES)
    duopoly_pairs = load_file_text(DUOPOLY_FILE)

    stocks_summary = json.dumps(stocks[:35], indent=2)

    system_prompt = """You are an expert technical analyst for Indian equity markets.
You have deep knowledge of:
- RSI patterns and thresholds for Indian mid/small-cap stocks
- Sector duopoly pairs in Indian markets
- Product launch and technology sector momentum
- Historical pattern performance in BSE/NSE

PATTERN RECOGNITION RULES:
a) Duopoly Pattern: Identify duopoly pairs (e.g., MRPL & CPCL, PVR & Inox, SRF & Navin Fluorine).
   Signal: If peer has risen but this stock has not yet reacted, flag as BUY candidate.

b) RSI Ceiling Rule: NEVER recommend stocks with RSI > 75.

c) RSI Recovery Pattern: Flag stocks that crossed RSI 45 from below RSI 30 in past 2 days.
   (Adjust threshold based on pattern_notes if available)

d) Strong Stock Dip Pattern: Stocks consistently above RSI 50 for 30+ days, briefly dipped,
   now recovering above RSI 45 — institutional strength indicator.

e) Product Launch Excitement: Stocks tied to recent product launches with genuine consumer excitement.

f) Trending Technology: EV, Solar, Defense Tech, AI/Data Center, Semiconductors, Space Tech sectors.

g) Self-Discovered: Any new patterns with strong rationale.

CONSTRAINTS:
- Maximum 7 recommendations
- RSI range: 35-68 preferred (never > 75)
- Consider current market conditions: India markets in May 2026
- Weight patterns by historical accuracy from pattern_notes

Return a JSON array (no markdown) with up to 7 entries:
[{
  "symbol": "NSE_SYMBOL",
  "company_name": "Full Name",
  "patterns_matched": ["pattern1", "pattern2"],
  "rsi": 52.3,
  "entry_price": 456.0,
  "target_price": 489.0,
  "stop_loss": 440.0,
  "confidence_score": 82,
  "reason": "Detailed 3-4 sentence reasoning covering technicals, fundamentals, and pattern match"
}]

After your JSON, on a new line starting with "PATTERN_NOTES_UPDATE:", write a brief update
for pattern_notes.md (what patterns were applied, observations, confidence levels)."""

    user_msg = f"""Date: {TODAY_STR} (Thursday — pre-weekend trade)

SCREENED STOCKS:
{stocks_summary}

EXISTING PATTERN NOTES:
{pattern_notes if pattern_notes else "No prior notes — first run."}

KNOWN DUOPOLY PAIRS:
{duopoly_pairs if duopoly_pairs else "No cached pairs — discover fresh."}

Market Context May 8, 2026:
- India VIX likely in 13-16 range (estimate)
- FII flows mixed but domestic SIP flows strong
- Budget cycle complete; Q4 FY26 results season ongoing
- Sectors in focus: Defense, Railway, Power, Capital Goods, IT services recovery
- Global context: US Fed on hold, geopolitical tensions moderate
- INR stable around 83-84 per USD

Analyze the stocks and return your top 7 recommendations as JSON, then add PATTERN_NOTES_UPDATE:"""

    print("  Calling Opus for pattern analysis (this may take 30-60s)...")
    raw = call_claude("analyzer", "claude-opus-4-7", system_prompt, user_msg, max_tokens=6000)

    # Split JSON from notes update
    pattern_update = ""
    json_part = raw
    if "PATTERN_NOTES_UPDATE:" in raw:
        parts = raw.split("PATTERN_NOTES_UPDATE:", 1)
        json_part = parts[0].strip()
        pattern_update = parts[1].strip()

    # Parse recommendations
    try:
        json_clean = json_part.strip()
        if json_clean.startswith("```"):
            json_clean = json_clean.split("```")[1]
            if json_clean.startswith("json"):
                json_clean = json_clean[4:]
        recommendations = json.loads(json_clean)
    except Exception as e:
        print(f"  [WARN] JSON parse error: {e}. Attempting repair...")
        # Try to find JSON array
        import re
        match = re.search(r'\[.*\]', json_part, re.DOTALL)
        if match:
            try:
                recommendations = json.loads(match.group())
            except:
                recommendations = []
        else:
            recommendations = []

    # Save pattern notes
    if pattern_update:
        existing = load_file_text(PATTERN_NOTES)
        updated = f"# Pattern Notes — Updated {TODAY_STR}\n\n{pattern_update}\n\n---\n\n{existing}"
        PATTERN_NOTES.write_text(updated)
        print(f"  Pattern notes updated.")

    # Save/update duopoly pairs from recommendations
    duopoly_from_analysis = []
    for r in recommendations:
        if "duopoly" in [p.lower() for p in r.get("patterns_matched", [])]:
            duopoly_from_analysis.append(r.get("symbol", ""))

    if duopoly_from_analysis:
        existing_pairs = json.loads(duopoly_pairs) if duopoly_pairs else {}
        existing_pairs[TODAY_STR] = duopoly_from_analysis
        DUOPOLY_FILE.write_text(json.dumps(existing_pairs, indent=2))

    print(f"  Generated {len(recommendations)} recommendations.")
    return recommendations[:7]


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 – VALIDATION (Sonnet)
# ═══════════════════════════════════════════════════════════════════════════════
def step3_validator(recommendations: list) -> list:
    print("\n=== STEP 3: VALIDATION (Sonnet) ===")

    rec_json = json.dumps(recommendations, indent=2)

    system_prompt = """You are a risk validation agent for Indian stock market recommendations.
For each stock, perform these validation checks:

a) NEGATIVE NEWS CHECK: Search your knowledge for any recent (last 7 days as of May 8, 2026) negative news:
   - Regulatory/SEBI/CCI issues, fraud allegations, management changes, earnings misses, legal troubles
   - If significant negative news: mark negative_news=true with reason

b) US MARKET STATUS: Assess last US market session (May 7, 2026):
   - If NASDAQ/S&P 500 closed DOWN >1%: us_market_status="negative", add US_MARKET_CAUTION
   - If DOWN 0-1%: us_market_status="neutral"
   - If UP: us_market_status="positive"

c) INDUSTRY TREND CHECK:
   - "growing": sector has expanding revenues, regulatory tailwinds, structural demand
   - "stable": steady business, no major headwinds
   - "fading": declining revenues, regulatory headwinds, obsolescence risk
   - Remove stocks with "fading" industry trend

d) GOLD PRICE CHECK: If gold has risen >2% in last 5 days (equity stress signal):
   - gold_caution=true for ALL recommendations
   - As of May 8, 2026, assess gold trend from your knowledge

DECISION RULES:
- negative_news=true → final_recommendation=false (strike the stock)
- industry_trend="fading" → final_recommendation=false
- All others: final_recommendation=true (with any caution flags noted)

Return a JSON array (no markdown) — one object per stock:
[{
  "symbol": "SYMBOL",
  "company_name": "Name",
  "negative_news": false,
  "negative_news_reason": "",
  "us_market_status": "positive|negative|neutral",
  "industry_trend": "growing|stable|fading",
  "industry_trend_reason": "brief explanation",
  "gold_caution": false,
  "us_market_caution": false,
  "final_recommendation": true,
  "validation_note": "Any additional risk notes"
}]"""

    user_msg = f"""Date: {TODAY_STR}

RECOMMENDATIONS TO VALIDATE:
{rec_json}

Validate each recommendation carefully. Consider:
- Indian market context: May 8, 2026 (Q4 results season, post-election stability)
- US markets: Assess May 7, 2026 close direction
- Gold prices: Has gold been rising the past 5 trading days?
- Any recent corporate governance issues at these companies

Return validation JSON array only (no markdown fences)."""

    print("  Calling Sonnet for validation...")
    raw = call_claude("validator", "claude-sonnet-4-6", system_prompt, user_msg, max_tokens=4096)

    try:
        raw_clean = raw.strip()
        if raw_clean.startswith("```"):
            raw_clean = raw_clean.split("```")[1]
            if raw_clean.startswith("json"):
                raw_clean = raw_clean[4:]
        validated = json.loads(raw_clean)
    except Exception as e:
        print(f"  [WARN] Validation JSON parse error: {e}. Using pass-through.")
        validated = [
            {
                "symbol": r.get("symbol"),
                "company_name": r.get("company_name"),
                "negative_news": False,
                "negative_news_reason": "",
                "us_market_status": "neutral",
                "industry_trend": "stable",
                "industry_trend_reason": "Defaulted due to parse error",
                "gold_caution": False,
                "us_market_caution": False,
                "final_recommendation": True,
                "validation_note": "Validation parse error — manual check recommended"
            }
            for r in recommendations
        ]

    # Merge validation data into recommendations
    validation_map = {v["symbol"]: v for v in validated}
    merged = []
    for r in recommendations:
        sym = r.get("symbol", "")
        v = validation_map.get(sym, {})
        merged.append({**r, **v})

    passed = [m for m in merged if m.get("final_recommendation", True)]
    failed = [m for m in merged if not m.get("final_recommendation", True)]

    print(f"  Validation: {len(passed)} passed, {len(failed)} struck out")
    for f in failed:
        print(f"    STRUCK: {f['symbol']} — {f.get('negative_news_reason') or f.get('industry_trend_reason', 'validation failed')}")

    return merged


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 – RESULT FORMATTING (Haiku)
# ═══════════════════════════════════════════════════════════════════════════════
def step4_formatter(merged: list, all_stocks: list) -> str:
    print("\n=== STEP 4: RESULT FORMATTING (Haiku) ===")

    # Filter to final recommendations, top 5
    final = [m for m in merged if m.get("final_recommendation", True)]
    final_sorted = sorted(final, key=lambda x: (-x.get("confidence_score", 0),
                                                  -x.get("avg_daily_turnover_lakh", 0)))[:5]

    # Build stock details for formatter
    stock_lookup = {s.get("symbol"): s for s in all_stocks}

    enriched = []
    for r in final_sorted:
        sym = r.get("symbol", "")
        base = stock_lookup.get(sym, {})
        enriched.append({**base, **r})

    enriched_json = json.dumps(enriched, indent=2)

    # Calculate token costs (approximate INR conversion: $1 = ₹83.5)
    USD_TO_INR = 83.5
    # Haiku: $1/$5 per 1M input/output
    # Opus 4.7: $5/$25 per 1M input/output
    # Sonnet 4.6: $3/$15 per 1M input/output
    costs = {
        "screener": (token_usage["screener"]["input"] * 1.0 + token_usage["screener"]["output"] * 5.0) / 1_000_000,
        "analyzer": (token_usage["analyzer"]["input"] * 5.0 + token_usage["analyzer"]["output"] * 25.0) / 1_000_000,
        "validator": (token_usage["validator"]["input"] * 3.0 + token_usage["validator"]["output"] * 15.0) / 1_000_000,
    }

    system_prompt = """You are a financial report formatter for Indian stock market recommendations.

Create a professional, detailed terminal/text report. Use ASCII formatting, borders, and charts.

FORMAT REQUIREMENTS:
1. Show top 5 stocks by confidence score (sorted: confidence DESC, volume DESC)
2. For each stock show:
   - Header with stock name, symbol, sector, cap category
   - Screening criteria checklist (✓/✗) for: Price>₹25, Market Cap>₹1000Cr, YoY Profit>25%, Daily Turnover>₹1Cr
   - Pattern checklist: which patterns matched
   - Validation status (Passed/Flagged/Struck)
   - Key metrics: P/E (estimate), RSI, Last Close, Avg Volume Turnover, Market Cap
   - Entry/Target/Stop-Loss prices
   - Potential upside % = (target - entry) / entry * 100
   - Risk-Reward ratio
   - 2-3 sentence investment thesis
   - ASCII sparkline of estimated 10-day price trend (use: ▁▂▃▄▅▆▇█ characters)
3. Market Caution Flags section (if any)
4. Disclaimer

Be thorough but readable. Use box-drawing characters (═╔╗╚╝║╠╣╦╩╬) for borders.
Do NOT use markdown (no # headers, no ** bold, no backticks).
Use ALL CAPS for section headers."""

    user_msg = f"""Date: {TODAY_STR}

FINAL RECOMMENDATIONS DATA:
{enriched_json}

TOKEN USAGE (will be appended separately):
Screener (Haiku):   {token_usage["screener"]["input"]} in / {token_usage["screener"]["output"]} out
Analyzer (Opus):    {token_usage["analyzer"]["input"]} in / {token_usage["analyzer"]["output"]} out
Validator (Sonnet): {token_usage["validator"]["input"]} in / {token_usage["validator"]["output"]} out

Generate the complete formatted report with ASCII charts, all checklists, and investment theses.
End the report before the token cost section — that will be appended separately."""

    print("  Calling Haiku for report formatting...")
    report = call_claude("formatter", "claude-haiku-4-5", system_prompt, user_msg, max_tokens=6000)

    # Append token cost section
    formatter_cost = (token_usage["formatter"]["input"] * 1.0 + token_usage["formatter"]["output"] * 5.0) / 1_000_000
    total_usd = sum(costs.values()) + formatter_cost
    total_inr = total_usd * USD_TO_INR

    token_report = f"""
╔══════════════════════════════════════════════════════════════════╗
║                     TOKEN USAGE & COST                           ║
╠══════════════════════════════════════════════════════════════════╣
║ Screener  (Haiku):    {token_usage["screener"]["input"]:>6} in / {token_usage["screener"]["output"]:>5} out  ≈ ${costs["screener"]:.4f} USD  ║
║ Analyzer  (Opus 4.7): {token_usage["analyzer"]["input"]:>6} in / {token_usage["analyzer"]["output"]:>5} out  ≈ ${costs["analyzer"]:.4f} USD  ║
║ Validator (Sonnet):   {token_usage["validator"]["input"]:>6} in / {token_usage["validator"]["output"]:>5} out  ≈ ${costs["validator"]:.4f} USD  ║
║ Formatter (Haiku):    {token_usage["formatter"]["input"]:>6} in / {token_usage["formatter"]["output"]:>5} out  ≈ ${formatter_cost:.4f} USD  ║
╠══════════════════════════════════════════════════════════════════╣
║ TOTAL COST:  ${total_usd:.4f} USD  ≈  ₹{total_inr:.2f}                    ║
╚══════════════════════════════════════════════════════════════════╝
"""
    return report + token_report


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5 – BACKTEST & PERFORMANCE REPORT
# ═══════════════════════════════════════════════════════════════════════════════
def step5_backtest(final_recommendations: list) -> str:
    print("\n=== STEP 5: BACKTEST & PERFORMANCE REPORT ===")

    # Load existing daily recommendations
    existing_recs = []
    if DAILY_REC_FILE.exists():
        try:
            existing_recs = json.loads(DAILY_REC_FILE.read_text())
        except:
            existing_recs = []

    # Add today's recommendations
    today_entry = {
        "date": TODAY_STR,
        "recommendations": [r.get("symbol") for r in final_recommendations if r.get("final_recommendation")],
        "entry_prices": {r.get("symbol"): r.get("entry_price", r.get("last_close", 0))
                         for r in final_recommendations if r.get("final_recommendation")}
    }
    existing_recs.append(today_entry)
    DAILY_REC_FILE.write_text(json.dumps(existing_recs, indent=2))

    # Get last 21 days of data for backtest
    recent_recs = [r for r in existing_recs
                   if r.get("date", "") >= "2026-04-17"]  # ~21 days back

    if len(recent_recs) <= 1:
        return """
╔══════════════════════════════════════════════════════════════════╗
║           BACKTEST REPORT: LAST 3 WEEKS                          ║
╠══════════════════════════════════════════════════════════════════╣
║  FIRST RUN — No historical data available for backtesting.       ║
║  Today's recommendations have been logged to:                    ║
║  daily_recommendations.json                                      ║
║                                                                  ║
║  After 2+ trading days, run again to see performance tracking.   ║
╚══════════════════════════════════════════════════════════════════╝
"""

    # Build backtest prompt for prior recommendations (exclude today)
    prior_recs = [r for r in recent_recs if r.get("date") != TODAY_STR]
    recs_json = json.dumps(prior_recs[-15:], indent=2)  # last 15 entries max

    system_prompt = """You are a trade performance analyst for Indian stock markets.

Given historical recommendations, calculate hypothetical portfolio performance:
- Assume ₹10,000 invested equally across all recommended stocks on buy date
- Exit after 2 trading days (T+2 close)
- Use your knowledge of NSE closing prices for the given dates

For each recommendation date, estimate:
- Buy price (use the entry_price from recommendations)
- Sell price at T+2 (estimate based on your knowledge of stock movements)
- P&L = sum of ((sell - buy) / buy * investment_per_stock)
- Return% = P&L / total_invested * 100

Output a formatted ASCII table followed by summary statistics.
Use box-drawing characters. Be realistic with price estimates.
Format:
Date | Stocks | Avg Buy | Avg Sell T+2 | P&L (₹) | Return%
Then: TOTAL, AVG/DAY, BEST DAY, WORST DAY
Also: Total capital deployed, total profit, accuracy (% profitable days)"""

    user_msg = f"""Historical recommendations to backtest:
{recs_json}

Today's date: {TODAY_STR}
Today's recommendations have been logged but NOT included in backtest (no T+2 data yet).

Calculate performance for available historical dates and display the backtest table.
Use realistic NSE price estimates based on your training knowledge."""

    raw_backtest = call_claude("formatter", "claude-haiku-4-5", system_prompt, user_msg, max_tokens=3000)

    return raw_backtest


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 70)
    print(f"  INDIAN STOCK MARKET RECOMMENDATION PIPELINE")
    print(f"  Date: {TODAY_STR} | Powered by Claude (Haiku + Opus + Sonnet)")
    print("=" * 70)

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("\n[ERROR] ANTHROPIC_API_KEY environment variable not set.")
        sys.exit(1)

    errors = []

    # Step 1: Screen
    try:
        screen_result = step1_screener()
        all_stocks = screen_result["stocks"]
        print(f"  Screened {len(all_stocks)} stocks.")
    except Exception as e:
        print(f"  [ERROR] Screener failed: {e}")
        errors.append(f"Step 1 error: {e}")
        all_stocks = _fallback_stocks()

    # Step 2: Analyze
    try:
        recommendations = step2_analyzer(all_stocks)
    except Exception as e:
        print(f"  [ERROR] Analyzer failed: {e}")
        errors.append(f"Step 2 error: {e}")
        recommendations = []

    if not recommendations:
        print("  [WARN] No recommendations generated. Pipeline incomplete.")
        return

    # Step 3: Validate
    try:
        validated = step3_validator(recommendations)
    except Exception as e:
        print(f"  [ERROR] Validator failed: {e}")
        errors.append(f"Step 3 error: {e}")
        validated = recommendations  # pass through on error

    # Step 4: Format
    try:
        report = step4_formatter(validated, all_stocks)
    except Exception as e:
        print(f"  [ERROR] Formatter failed: {e}")
        errors.append(f"Step 4 error: {e}")
        report = f"Formatting error: {e}\n\nRaw recommendations:\n{json.dumps(validated, indent=2)}"

    # Step 5: Backtest
    final_rec_list = [v for v in validated if v.get("final_recommendation", True)]
    try:
        backtest_report = step5_backtest(final_rec_list)
    except Exception as e:
        print(f"  [ERROR] Backtest failed: {e}")
        errors.append(f"Step 5 error: {e}")
        backtest_report = f"Backtest error: {e}"

    # ── FINAL OUTPUT ──
    print("\n")
    print("=" * 70)
    print(report)
    print("\n")
    print("=" * 70)
    print("  BACKTEST & PERFORMANCE TRACKING")
    print("=" * 70)
    print(backtest_report)

    if errors:
        print("\n[PIPELINE WARNINGS]")
        for err in errors:
            print(f"  - {err}")

    print("\n[DISCLAIMER] This system is for informational and research purposes only.")
    print("Always validate recommendations with your own research before investing.")
    print("Past performance of patterns does not guarantee future results.")


if __name__ == "__main__":
    main()
